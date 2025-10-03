import os
import re
import zipfile
import base64
import requests
import tempfile
import time
from io import BytesIO
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook, load_workbook
from PIL import Image
from dotenv import load_dotenv

# ========= Env & API =========
load_dotenv()
API_KEY = os.getenv("OPENAI_API_KEY", "")
EXCEL_METADATA_PATH = os.getenv("EXCEL_METADATA_PATH", "")

API_URL = "https://api.openai.com/v1/responses"
MODEL = "gpt-4o-mini"  # oder "gpt-4o"

print("=== Museum Captioner — Excel context, AI-only descriptions ===")

# ========= Helpers =========
CODE_RX = re.compile(r'(\d+)[_\-\/](\d{4})[_\-\/](\d{3,4})(?:[_\-\/]\d{1,4})*')

def _pad4(x: str) -> str:
    return str(x).zfill(4)

def normalize_code(b1: str, b2: str, b3: str) -> str:
    return f"{int(b1)}-{b2}-{_pad4(b3)}"

def normalize_item_id(s: str) -> str:
    if not isinstance(s, str):
        s = str(s or "")
    s = s.strip()
    m = CODE_RX.search(s) or CODE_RX.search(os.path.splitext(os.path.basename(s))[0])
    return normalize_code(*m.groups()) if m else ""

def code_variants(code: str) -> set:
    out = set()
    m = re.match(r'^(\d+)-(\d{4})-(\d{3,4})$', code)
    if not m:
        return {code} if code else set()
    b1, y, n = m.groups()
    out.add(f"{int(b1)}-{y}-{_pad4(n)}")  # 0736
    out.add(f"{int(b1)}-{y}-{int(n)}")    # 736
    return out

def is_real_image_file(name: str) -> bool:
    n = name.lower()
    if n.startswith("._") or n in {".ds_store"}:
        return False
    return n.endswith((".png", ".jpg", ".jpeg", ".webp"))

def encode_image(image_path: str) -> str:
    with Image.open(image_path) as im:
        im = im.convert("RGB")
        buf = BytesIO()
        ext = (os.path.splitext(image_path)[1] or "").lower()
        if ext == ".png":
            im.save(buf, format="PNG", optimize=True)
            mime = "image/png"
        else:
            im.save(buf, format="JPEG", quality=90, optimize=True)
            mime = "image/jpeg"
        b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
        return f"data:{mime};base64,{b64}"

def extract_item_id(filename: str) -> str:
    return normalize_item_id(os.path.basename(filename))

# ========= Excel Loader =========
def find_codes_in_row(row_values) -> set:
    codes = set()
    for val in row_values:
        if val is None: 
            continue
        for m in CODE_RX.finditer(str(val)):
            codes.add(normalize_code(*m.groups()))
    return codes

def load_excel_descriptions(path: str) -> dict:
    if not path:
        raise FileNotFoundError("EXCEL_METADATA_PATH fehlt (.env).")
    p = Path(path).expanduser()
    if not p.is_file():
        raise FileNotFoundError(f"Excel nicht gefunden: {p}")

    wb = load_workbook(str(p), data_only=True)
    ws = wb.active
    mapping = {}

    # Aus deinem Sheet: F=5 (maker), BV=73 (measurements), CD=82 (date)
    MAKER_COL_IDX = 5
    MEAS_COL_IDX  = 73
    DATE_COL_IDX  = 82

    total_rows = 0
    total_links = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        maker = str((row[MAKER_COL_IDX] if len(row) > MAKER_COL_IDX else "") or "").strip()
        meas  = str((row[MEAS_COL_IDX]  if len(row) > MEAS_COL_IDX  else "") or "").strip()
        date  = str((row[DATE_COL_IDX]  if len(row) > DATE_COL_IDX  else "") or "").strip()

        codes = find_codes_in_row(row)
        if not codes:
            continue

        rec = {"maker": maker, "date": date, "measurements": meas}
        for c in codes:
            for v in code_variants(c):
                mapping[v] = rec
                total_links += 1

    print(f"[META] Scanned {total_rows} rows -> built {total_links} keys ({len(set(mapping.keys()))} unique).")
    # Debug-Beispiele
    for i, (k, v) in enumerate(mapping.items()):
        print(f"[META] EXAMPLE -> {k}: {v}")
        if i >= 1: break
    return mapping

# ========= Prompt & KI =========
def build_prompt_de(excel_info: dict | None) -> str:
    """
    Baut den strikten Prompt mit Regeln + optionalem Kontext aus Excel.
    """
    base = """
You are a museum conservator for the Deutsches Technikmuseum. Your task is to create an **objective** description of the depicted object(s) based **only** on what is visibly present in the image.

Objective & Scope (Mandatory)
- Describe only physical characteristics that are clearly visible (form, visually evident materials, color/finish, construction features, condition traces, notable markings/text).
- Exclude backgrounds, stands, rulers, mounts, hands, labels, and studio setup.
- Do not infer function, date, origin, brand/model names, maker, usage, or history unless **clearly shown** (e.g., a readable stamped logo).
- If a detail is not clearly visible, do not mention it.

Style
- Neutral, factual tone in **English**.
- No promotional language or superlatives; no catalog clichés.
- Length may vary with object complexity (prefer 1–3 compact sentences; do not force a count).

Use of Metadata (if present below)
- Use metadata **only** when consistent with what is visible. If a metadata claim is not visually supported, omit it from the description.
- Never invent or complete missing metadata. Never output exact measurements or weights.

Ambiguity & Multi-object Scenes
- If multiple distinct objects are present, focus on the primary object (largest or centered).
- If the primary object cannot be determined, add the flag `primary_object_unclear` and keep the text generic.

Quality & Uncertainty
- If the image is too small, blurred, or cropped to judge key features, add the flag `low_image_quality` and keep the description minimal.
- Avoid guesses.

Prohibitions (Mandatory)
- No brands/model names unless clearly visible.
- No advertising language.
- No interpretation (purpose, era, provenance).
- No exact dimensions/weights.
- Do not reproduce existing catalog text verbatim.
"""

    # optional: Kontext aus Excel, aber klar getrennt
    if excel_info:
        ctx = []
        if excel_info.get("maker"):
            ctx.append(f"Hersteller/Urheber (Metadaten): {excel_info['maker']}")
        if excel_info.get("date"):
            ctx.append(f"Datierung (Metadaten): {excel_info['date']}")
        if excel_info.get("measurements"):
            ctx.append("Hinweis: Maße sind im Katalog vorhanden, dürfen aber nicht wörtlich ausgegeben werden.")
        if ctx:
            base += "\n\nMetadaten (nur verwenden, falls im Bild visuell bestätigt):\n" + "\n".join(ctx)

    return base

def get_caption(encoded_image_url: str, excel_info: dict | None = None) -> str:
    if not API_KEY:
        return "OPENAI_API_KEY fehlt (.env)."
    prompt = build_prompt_de(excel_info)
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {API_KEY}"}
    payload = {
        "model": MODEL,
        "input": [{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_image", "image_url": encoded_image_url},
            ],
        }],
        "temperature": 0.2,
        "max_output_tokens": 220,
    }

    for attempt in range(5):
        try:
            r = requests.post(API_URL, headers=headers, json=payload, timeout=120)
            if r.status_code == 200:
                data = r.json()
                outputs = data.get("output", [])
                if outputs and outputs[0].get("content"):
                    return outputs[0]["content"][0].get("text", "").strip()
                return "Fehler: leere KI-Antwort"
            elif r.status_code in (429, 500, 502, 503, 504):
                sleep = 0.5 * (2 ** attempt)
                print(f"[INFO] Backing off {sleep:.1f}s (HTTP {r.status_code})")
                time.sleep(sleep)
                continue
            else:
                print(f"[WARN] API error {r.status_code}: {r.text}")
                return "Fehler bei der KI-Anfrage"
        except Exception as e:
            print(f"[WARN] Unexpected error: {e}")
            time.sleep(0.5 * (2 ** attempt))
    return "Fehler bei der KI-Anfrage"

# ========= Verarbeitung =========
def lookup_excel(mapping: dict, item_id: str) -> dict | None:
    norm = normalize_item_id(item_id)
    for k in code_variants(norm):
        if k in mapping:
            return mapping[k]
    return None

def process_images(input_path: str, output_excel: str, excel_map: dict):
    with tempfile.TemporaryDirectory() as tmp:
        directory_to_process = tmp if zipfile.is_zipfile(input_path) else input_path
        if zipfile.is_zipfile(input_path):
            with zipfile.ZipFile(input_path, "r") as z:
                z.extractall(tmp)

        grouped = defaultdict(list)
        for root, _, files in os.walk(directory_to_process):
            for file_name in sorted(files):
                if not is_real_image_file(file_name):
                    continue
                item_id = extract_item_id(file_name)
                if item_id:
                    grouped[item_id].append(os.path.join(root, file_name))

        wb = Workbook()
        ws = wb.active
        ws.title = "Descriptions"
        ws.append(["item_id", "maker", "date", "measurements", "final_description", "source", "context_used"])

        for item_id, image_paths in grouped.items():
            key_norm = normalize_item_id(item_id)
            excel_info = lookup_excel(excel_map, item_id)
            maker = excel_info.get("maker", "") if excel_info else ""
            date = excel_info.get("date", "") if excel_info else ""
            measurements = excel_info.get("measurements", "") if excel_info else ""
            context_used = "yes" if excel_info else "no"

            picked = next((p for p in image_paths if os.path.exists(p)), None)
            if picked:
                encoded = encode_image(picked)
                ai_desc = get_caption(encoded, excel_info)
                print(f"[AI] {key_norm} -> Generated (context_used={context_used})")
            else:
                ai_desc = "Kein Bild gefunden"

            # WICHTIG: final_description ist IMMER die KI-Ausgabe
            ws.append([key_norm, maker, date, measurements, ai_desc, "ai", context_used])
            time.sleep(0.1)

        wb.save(output_excel)
        print(f"[DONE] Excel saved: {output_excel}")

# ========= Entry =========
def main():
    input_path = input("Select path to your ZIP file or folder: ").strip()
    output_excel = "descriptions_with_excel.xlsx"

    if not API_KEY:
        raise ValueError("API key missing (.env OPENAI_API_KEY).")

    excel_map = load_excel_descriptions(EXCEL_METADATA_PATH)
    process_images(input_path, output_excel, excel_map)

if __name__ == "__main__":
    main()
